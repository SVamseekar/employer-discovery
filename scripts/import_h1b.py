# scripts/import_h1b.py
import openpyxl
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))
from schema import FIELDS, UNKNOWN, empty_row

XLSX_PATH = "data/raw/H1B_Companies_By_Sector.xlsx"
OUTPUT_PATH = "data/master_employers.csv"

SKIP_SHEETS = {"Overview"}

SECTOR_MAP = {
    "Financial Services & Fintech": ("FinTech", "Financial Services"),
    "Unexpected Tech": ("Enterprise SaaS", "Non-obvious Tech"),
    "Retail & E-Commerce": ("Commerce Infrastructure", "Retail"),
    "Healthcare & Health Tech": ("HealthTech", "Healthcare"),
    "Consulting & IT Services": ("Enterprise SaaS", "IT Consulting"),
    "Core Tech": ("AI Infrastructure", "Core Technology"),
    "Logistics & Supply Chain": ("Supply Chain", "Logistics"),
    "Manufacturing & Industrial": ("Industrial AI", "Manufacturing"),
    "Energy & Utilities": ("ClimateTech", "Energy & Utilities"),
    "Media & Telecom": ("Developer Infrastructure", "Media & Telecom"),
    "Food & Consumer Goods": ("Commerce Infrastructure", "Consumer Goods"),
    "Real Estate & PropTech": ("PropTech", "Real Estate"),
}

def import_h1b():
    wb = openpyxl.load_workbook(XLSX_PATH)
    rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        sector, subsector = SECTOR_MAP.get(sheet_name, ("Enterprise SaaS", sheet_name))

        for row in ws.iter_rows(min_row=4, values_only=True):
            company = row[0]
            if not company or not isinstance(company, str) or company == "Company":
                continue

            careers_url = row[9] if len(row) > 9 and row[9] else UNKNOWN
            notes = row[4] if len(row) > 4 and row[4] else ""
            state = row[2] if len(row) > 2 and row[2] else UNKNOWN

            r = empty_row()
            r["Company"] = company.strip()
            r["Website"] = UNKNOWN
            r["Careers_URL"] = careers_url if careers_url else UNKNOWN
            r["Country"] = "USA"
            r["City"] = f"{state}, USA" if state != UNKNOWN else "USA"
            r["Sector"] = sector
            r["Subsector"] = subsector
            r["Company_Stage"] = "Enterprise"
            r["Company_Scale"] = "1000+"
            r["Employer_Category"] = "H1B Sponsor"
            r["Remote"] = UNKNOWN
            r["Visa_Sponsorship"] = "Yes"
            r["EOR"] = UNKNOWN
            r["Hiring_Geography"] = "USA"
            r["Target_Roles"] = "Data Engineer / AI Engineer / Backend Engineer"
            r["Tech_Stack"] = UNKNOWN
            r["Region_Eligibility"] = "USA"
            r["Portfolio_Theme_Match"] = UNKNOWN
            r["Language_Requirement"] = "None"
            r["Hiring_Confidence"] = "Medium"
            r["Reason_Match"] = notes.strip() if notes else "Confirmed H1B sponsor"
            r["Source"] = "H1B_Companies_By_Sector.xlsx"
            r["Cold_Outreach_Candidate"] = "No"
            r["Visa_Sponsor_Register"] = "Unknown"
            rows.append(r)

    write_master(rows)
    print(f"Imported {len(rows)} companies from H1B xlsx → {OUTPUT_PATH}")

def write_master(rows):
    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)

if __name__ == "__main__":
    import_h1b()
